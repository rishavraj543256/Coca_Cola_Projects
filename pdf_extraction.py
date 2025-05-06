import pdfplumber
import re
import pandas as pd
import numpy as np
import os
from datetime import datetime
import warnings
import contextlib
import logging
import glob
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Configure pandas display settings
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', None)
pd.set_option('display.expand_frame_repr', False)


@contextlib.contextmanager
def suppress_warnings():
    """Context manager to suppress warnings and logging messages."""
    logging.getLogger("pdfminer").setLevel(logging.ERROR)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        yield


def is_summary_section(text):
    """
    Check if the given text is from the summary section.
    Returns tuple of (is_summary, section_end_index) where section_end_index is the line
    number where the summary section ends (or -1 if not found)
    """
    lines = text.split('\n')
    summary_start = -1
    summary_end = -1
    
    # Find the summary section boundaries
    for i, line in enumerate(lines):
        if 'Summary of information gathered' in line:
            summary_start = i
        elif summary_start != -1 and any(x in line for x in ['Annexure', 'Exhibit']):
            summary_end = i
            break
    
    if summary_start != -1:
        if summary_end == -1:  # If no Annexure found, consider rest of the text as summary
            summary_end = len(lines)
        # Return the text between summary start and end
        return True, '\n'.join(lines[summary_start:summary_end])
    
    return False, ''


def extract_table_data(page):
    """
    Extract manufacturer statistics table from a page.
    
    Args:
        page: pdfplumber page object
        
    Returns:
        list: List of dictionaries containing table data
    """
    def is_superscript(char):
        """Helper function to check if a character is a superscript number"""
        superscripts = '⁰¹²³⁴⁵⁶⁷⁸⁹'
        return char in superscripts
    
    def extract_base_number(text, header=''):
        """Extract the base number from a string that might contain exponential notation"""
        # Convert to string and strip whitespace
        text = str(text).strip()
        has_exponential = False
        
        # If empty, return '0'
        if not text:
            return '0', has_exponential
        
        # Split the text into characters
        chars = list(text)
        exp_start = -1
        
        # Look for superscript numbers
        for i, c in enumerate(chars):
            if is_superscript(c):
                exp_start = i
                has_exponential = True
                break
        
        # If no superscript found, look for ** pattern
        if exp_start == -1:
            for i in range(len(chars)-1):
                if chars[i] == '*' and chars[i+1] == '*':
                    exp_start = i
                    has_exponential = True
                    break
        
        # Check for potential converted superscript (ending in 6)
        if not has_exponential and 'repeat' in str(header).lower():
            text_str = str(text)
            if text_str.endswith('6'):
                has_exponential = True
        
        # If we found an exponential marker, only take the part before it
        if exp_start != -1:
            chars = chars[:exp_start]
        
        # Extract only digits and commas from the remaining characters
        base_number = ''.join(c for c in chars if c.isdigit() or c == ',')
        
        return base_number if base_number else '0', has_exponential
    
    # First check if this page contains the summary section
    text = page.extract_text()
    is_summary, summary_text = is_summary_section(text)
    if not is_summary:
        return []
    
    tables = page.extract_tables()
    manufacturer_data = []
    
    for table in tables:
        # Skip empty tables
        if not table:
            continue
        
        # Try to identify if this is a manufacturer statistics table
        header_row = [str(cell).strip() if cell else '' for cell in table[0]]
        
        # Check if this is a manufacturer table by looking for key column names
        if any('manufacturer' in str(cell).lower() for cell in header_row):
            # Process each row after header
            for row in table[1:]:
                # Skip empty rows or total rows
                if not row or all(not cell for cell in row) or (row[0] and 'total' in str(row[0]).lower()):
                    continue
                
                # Create a dictionary for this manufacturer
                manufacturer = {
                    'manufacturer_name': str(row[0]).strip() if row[0] else None,
                    'stock_observation_percentage': None,
                    'affected_loose_units': None,
                    'affected_full_cases': None,
                    'affected_loose_repeat_batch': None,
                    'affected_cases_repeat_batch': None,
                    'has_exponential': 'No'  # New field to flag exponential numbers
                }
                
                # Extract percentage from the row
                for cell in row:
                    if cell and '%' in str(cell):
                        try:
                            # Remove any non-numeric characters except decimal point
                            clean_value = ''.join(c for c in str(cell) if c.isdigit() or c == '.')
                            manufacturer['stock_observation_percentage'] = float(clean_value)
                        except ValueError:
                            pass
                
                # Extract numeric values for units and cases
                for i, cell in enumerate(row):
                    if not cell:
                        continue
                    
                    try:
                        # Extract the base number from the cell
                        clean_value, has_exp = extract_base_number(cell, header_row[i])
                        
                        if has_exp:
                            manufacturer['has_exponential'] = 'Yes - Check Values'
                        
                        if clean_value:  # Only try to convert if we have digits
                            value = int(clean_value.replace(',', ''))
                            
                            # Determine which field this number belongs to based on column header
                            header = str(header_row[i]).lower()
                            if 'loose' in header and 'repeat' in header:
                                manufacturer['affected_loose_repeat_batch'] = value
                            elif 'case' in header and 'repeat' in header:
                                manufacturer['affected_cases_repeat_batch'] = value
                            elif 'loose' in header:
                                manufacturer['affected_loose_units'] = value
                            elif 'case' in header:
                                manufacturer['affected_full_cases'] = value
                    except ValueError:
                        continue
                
                # Only add non-empty manufacturers
                if manufacturer['manufacturer_name'] and not any(x in str(manufacturer['manufacturer_name']).lower() for x in ['total', 'none']):
                    manufacturer_data.append(manufacturer)
    
    return manufacturer_data


def extract_summary_date(text):
    """Extract the date from the summary section."""
    # First find the summary section
    summary_match = re.search(r'Summary of information gathered[:\s\n]+(.*?)(?=\n\n|\Z)', text, re.DOTALL | re.IGNORECASE)
    if summary_match:
        summary_text = summary_match.group(1)
        
        # First try to find specifically the "received" date
        received_pattern = r'(?:request|mail)\s+was\s+received.*?on\s+(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})'
        received_match = re.search(received_pattern, summary_text, re.IGNORECASE)
        if received_match:
            return standardize_date_format(received_match.group(1))
        
        # If no received date found, look for other date patterns
        date_patterns = [
            r'(?:received|conducted).*?on\s+(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})',
            r'(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})'
        ]
        
        for pattern in date_patterns:
            date_match = re.search(pattern, summary_text)
            if date_match:
                return standardize_date_format(date_match.group(1))
        
        # If no standard date format found, try patterns with ordinal indicators
        ordinal_patterns = [
            r'(?:request|mail)\s+was\s+received.*?on\s+(\d{1,2})(?:st|nd|rd|th)\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}',
            r'(?:received|conducted).*?on\s+(\d{1,2})(?:st|nd|rd|th)\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}',
            r'(\d{1,2})(?:st|nd|rd|th)\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}'
        ]
        
        for pattern in ordinal_patterns:
            date_match = re.search(pattern, summary_text, re.IGNORECASE)
            if date_match:
                # Extract the full date and clean it
                full_date = date_match.group(0)
                # Remove ordinal indicators
                clean_date = re.sub(r'(\d{1,2})(?:st|nd|rd|th)', r'\1', full_date)
                # Extract just the date part if it was in a "received on" context
                if 'received' in clean_date or 'conducted' in clean_date:
                    date_only = re.search(r'(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})', clean_date)
                    if date_only:
                        clean_date = date_only.group(1)
                return standardize_date_format(clean_date)
    
    return None


def standardize_date_format(date_str):
    """
    Convert any recognized date format to DD-MM-YYYY format.
    
    Args:
        date_str: Date string in various formats
        
    Returns:
        str: Date in DD-MM-YYYY format or None if conversion fails
    """
    if pd.isna(date_str) or not date_str:
        return None
        
    try:
        # Convert month names to numbers using a custom mapping
        month_map = {
            'january': '01', 'jan': '01',
            'february': '02', 'feb': '02',
            'march': '03', 'mar': '03',
            'april': '04', 'apr': '04',
            'may': '05',
            'june': '06', 'jun': '06',
            'july': '07', 'jul': '07',
            'august': '08', 'aug': '08',
            'september': '09', 'sep': '09',
            'october': '10', 'oct': '10',
            'november': '11', 'nov': '11',
            'december': '12', 'dec': '12'
        }
        
        # Clean and standardize the input string
        date_str = str(date_str).strip().lower()
        
        # First try to match dates with month names (e.g., "02 April 2025")
        pattern = r'(\d{1,2})\s+([a-zA-Z]+)\s+(\d{4})'
        match = re.search(pattern, date_str)
        
        if match:
            day = match.group(1).zfill(2)  # Pad with leading zero if needed
            month_name = match.group(2).lower()[:3]  # Get first 3 chars of month name
            month = month_map.get(month_name)
            year = match.group(3)
            
            if month:  # If we successfully mapped the month name
                return f"{day}-{month}-{year}"
        
        # For dates already in DD-MM-YYYY format
        if re.match(r'^\d{2}-\d{2}-\d{4}$', date_str):
            return date_str
            
        # For other formats, try to parse with pandas but specify the format
        # This is mainly for backup, we should mostly catch dates in the above formats
        try:
            # First try parsing as DD-MM-YYYY
            return pd.to_datetime(date_str, format='%d-%m-%Y').strftime('%d-%m-%Y')
        except:
            try:
                # Then try as DD/MM/YYYY
                return pd.to_datetime(date_str, format='%d/%m/%Y').strftime('%d-%m-%Y')
            except:
                # Last resort - let pandas guess but force day first
                return pd.to_datetime(date_str, dayfirst=True).strftime('%d-%m-%Y')
    except:
        return None


def extract_pdf_info(pdf_path):
    """
    Extract specific information from the PDF file.
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        dict: Dictionary containing extracted information
    """
    # Check if file name contains required strings
    filename = os.path.basename(pdf_path).lower()
    if not ('draft report' in filename or 'draft findings' in filename):
        print(f"Skipping {pdf_path}: Filename does not contain 'Draft Report' or 'Draft Findings'")
        return {
            'file_name': os.path.basename(pdf_path),
            'company_name': None,
            'project_id': None,
            'location': None,
            'date': None,
            'survey_id': None,
            'requestor': None,
            'summary_date': None,
            'manufacturer_statistics': []
        }
    
    # Initialize dictionary to store extracted information
    info = {
        'file_name': os.path.basename(pdf_path),
        'company_name': None,
        'project_id': None,
        'location': None,
        'date': None,
        'survey_id': None,
        'requestor': None,
        'summary_date': None,  # New field for summary date
        'manufacturer_statistics': []
    }
    
    try:
        # Read the PDF file with suppressed warnings
        with suppress_warnings():
            with pdfplumber.open(pdf_path) as pdf:
                # Process first page for basic information
                if len(pdf.pages) > 0:
                    page = pdf.pages[0]
                    text = page.extract_text()
                    
                    if not text:
                        print(f"Warning: No text could be extracted from {pdf_path}")
                        return info
                    
                    # Extract summary date from first few pages
                    summary_date = None
                    for page_num in range(min(3, len(pdf.pages))):  # Check first 3 pages
                        page_text = pdf.pages[page_num].extract_text()
                        if page_text:
                            summary_date = extract_summary_date(page_text)
                            if summary_date:
                                info['summary_date'] = standardize_date_format(summary_date)
                                break
                    
                    # Try to detect PDF type based on content
                    is_draft_finding = any('Draft Finding' in line for line in text.split('\n'))
                    
                    if is_draft_finding:
                        # Extract Survey ID
                        for line in text.split('\n'):
                            if 'Survey ID:' in line or 'Survey ID' in line:
                                survey_match = re.search(r'(?:Survey ID:?\s*)?([A-Z]{2}\d{3,4})', line)
                                if survey_match:
                                    info['survey_id'] = survey_match.group(1)
                                    info['project_id'] = survey_match.group(1)  # Use survey ID as project ID
                                break
                        
                        # Extract Requestor
                        for i, line in enumerate(text.split('\n')):
                            if 'Requestor:' in line or 'Requestor' in line:
                                if i + 1 < len(text.split('\n')):
                                    requestor = text.split('\n')[i]
                                    if 'Requestor:' in requestor:
                                        requestor = requestor.split('Requestor:')[1]
                                    info['requestor'] = requestor.strip()
                                    info['company_name'] = requestor.strip()  # Use requestor as company name
                                break
                        
                        # Extract Location (City/State)
                        location_found = False
                        for line in text.split('\n'):
                            # First try to find lines with state names
                            if any(state in line for state in ['Pradesh', 'UP', 'Bihar', 'Maharashtra', 'Karnataka']):
                                # Split by comma and take everything before the state name
                                for state in ['Pradesh', 'UP', 'Bihar', 'Maharashtra', 'Karnataka']:
                                    if state in line:
                                        location = line[:line.index(state)].strip()
                                        if location.endswith(','):
                                            location = location[:-1].strip()
                                        info['location'] = location
                                        location_found = True
                                        break
                                if location_found:
                                    break
                            
                            # Look for Delhi variations and other city names
                            elif any(city in line for city in ['West Delhi', 'East Delhi', 'North Delhi', 'South Delhi', 'Central Delhi', 'Delhi', 'Mumbai', 'Kolkata', 'Bangalore', 'Hyderabad']):
                                # Try to match multi-word Delhi locations first
                                delhi_match = re.search(r'((?:West|East|North|South|Central)\s+Delhi|Delhi|Mumbai|Kolkata|Bangalore|Hyderabad)', line)
                                if delhi_match:
                                    info['location'] = delhi_match.group(1)
                                    location_found = True
                                    break
                        
                        # If location not found in the usual places, look in the header block
                        if not location_found:
                            header_text = '\n'.join(text.split('\n')[:20])  # Check first 20 lines for header info
                            location_lines = [line.strip() for line in header_text.split('\n') 
                                           if line.strip() and not any(x in line.lower() for x in 
                                           ['draft', 'finding', 'survey', 'id:', 'requestor:', 'date:', 'confidential'])]
                            
                            for line in location_lines:
                                if not any(x in line.lower() for x in ['ltd', 'limited', '@', 'email', 'tel', 'phone']):
                                    # Try to match multi-word Delhi locations first
                                    delhi_match = re.search(r'((?:West|East|North|South|Central)\s+Delhi|Delhi|Mumbai|Kolkata|Bangalore|Hyderabad)', line)
                                    if delhi_match:
                                        info['location'] = delhi_match.group(1)
                                        location_found = True
                                        break
                                    else:
                                        # If no Delhi match, try other cities/states
                                        for state in ['Pradesh', 'UP', 'Bihar', 'Maharashtra', 'Karnataka']:
                                            if state in line:
                                                location = line[:line.index(state)].strip()
                                                if location.endswith(','):
                                                    location = location[:-1].strip()
                                                info['location'] = location
                                                location_found = True
                                                break
                                if location_found:
                                    break
                        
                        # Extract Date
                        date_patterns = [
                            r'(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})',
                            r'(\d{2}(?:\s+|-|/)?(?:January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:\s+|-|/)\d{4})',
                            r'(\d{4}-\d{2}-\d{2})',
                            r'(\d{2}/\d{2}/\d{4})',
                            r'(\d{2}-[A-Za-z]{3}-\d{4})'
                        ]
                        
                        for pattern in date_patterns:
                            match = re.search(pattern, text)
                            if match:
                                info['date'] = standardize_date_format(match.group(1))
                                break
                    
                    else:
                        # Original PDF format processing
                        # Extract company name
                        project_stellar_index = -1
                        for i, line in enumerate(text.split('\n')):
                            if 'Project Stellar' in line:
                                project_stellar_index = i
                                break
                        
                        if project_stellar_index > 0:
                            company_lines = []
                            for i in range(max(0, project_stellar_index - 3), project_stellar_index):
                                line = text.split('\n')[i].strip()
                                if line and not any(x in line.lower() for x in ['confidential', 'draft report', 'report']):
                                    company_lines.append(line)
                            
                            if company_lines:
                                company_name = '\n'.join(company_lines)
                                info['company_name'] = clean_company_name(company_name)
                        
                        # Extract Project ID
                        id_patterns = [
                            r'(?:Project|ID|No|Number):\s*([A-Z]+\d+(?:\.[A-Z]\d*)?)',
                            r'(?:Project|ID|No|Number)\s+([A-Z]+\d+(?:\.[A-Z]\d*)?)',
                            r'([A-Z]{2,5}\d{2,5}(?:\.[A-Z]\d*)?)'
                        ]
                        
                        for pattern in id_patterns:
                            matches = re.finditer(pattern, text)
                            for match in matches:
                                potential_id = match.group(1)
                                if potential_id in os.path.basename(pdf_path):
                                    info['project_id'] = potential_id
                                    break
                            if info['project_id']:
                                break
                        
                        # Extract Location with support for multi-word cities
                        location_patterns = [
                            r'(?:Project|ID)[^,]*,\s*((?:West|East|North|South|Central)\s+Delhi|[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                            r'(?:[A-Z]+\d+)[^,]*,\s*((?:West|East|North|South|Central)\s+Delhi|[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                            r'Report_((?:West|East|North|South|Central)_Delhi|[A-Z][a-z]+(?:_[A-Z][a-z]+)*)',
                            r'((?:West|East|North|South|Central)\s+Delhi|[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)'
                        ]
                        
                        for pattern in location_patterns:
                            match = re.search(pattern, text)
                            if match:
                                location = match.group(1)
                                # Replace underscores with spaces if they exist
                                location = location.replace('_', ' ')
                                info['location'] = location.strip()
                                break
                        
                        # Extract Date
                        date_patterns = [
                            r'(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})',
                            r'(\d{2}(?:\s+|-|/)?(?:January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:\s+|-|/)\d{4})',
                            r'(\d{4}-\d{2}-\d{2})',
                            r'(\d{2}/\d{2}/\d{4})',
                            r'(\d{2}-[A-Za-z]{3}-\d{4})'
                        ]
                        
                        for pattern in date_patterns:
                            match = re.search(pattern, text)
                            if match:
                                info['date'] = standardize_date_format(match.group(1))
                                break
                
                # Process all pages for manufacturer statistics
                for page in pdf.pages:
                    manufacturer_data = extract_table_data(page)
                    if manufacturer_data:
                        info['manufacturer_statistics'].extend(manufacturer_data)
                
    except Exception as e:
        print(f"Error processing PDF {pdf_path}: {str(e)}")
    
    return info


def clean_company_name(name):
    """Clean up company name by removing extra whitespace only."""
    if not name:
        return name
    
    # Remove extra whitespace while preserving line breaks
    lines = [line.strip() for line in name.split('\n')]
    # Remove empty lines
    lines = [line for line in lines if line]
    # Join with newlines to preserve the multi-line format
    return '\n'.join(lines)


def select_file(title, file_types):
    """
    Open a file dialog for selecting a file.
    
    Args:
        title (str): Title of the file dialog window
        file_types (list): List of tuples containing file type descriptions and patterns
        
    Returns:
        str: Selected file path or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=file_types
    )
    root.destroy()
    return file_path if file_path else None


def process_all_pdfs_for_gui(directory, original_save_path=None, report_save_path=None):
    """
    Process all PDF files in the directory for GUI application.
    
    Args:
        directory (str): Directory containing PDF files
        original_save_path (str, optional): Path to save the original format Excel file
        report_save_path (str, optional): Path to save the report format Excel file
        
    Returns:
        tuple: (basic_df, manufacturer_df, original_path, report_path)
    """
    if not directory:
        print("No directory specified. Operation cancelled.")
        return None, None, None, None
    
    # Get all PDF files in the directory and subdirectories
    pdf_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                pdf_files.append(os.path.join(root, file))
    
    if not pdf_files:
        print(f"No PDF files found in {directory} or its subdirectories.")
        return None, None, None, None
    
    print(f"Found {len(pdf_files)} PDF files to process.")
    
    # Lists to store results
    basic_info_list = []
    manufacturer_stats_list = []
    
    for pdf_file in pdf_files:
        try:
            print(f"\nProcessing: {pdf_file}")
            info = extract_pdf_info(pdf_file)
            
            # Skip if the file doesn't match our criteria (checked in extract_pdf_info)
            if not info['manufacturer_statistics'] and info['company_name'] is None:
                continue
                
            # Add basic info
            basic_info = {
                'file_name': os.path.basename(pdf_file),
                'company_name': info.get('company_name', ''),
                'project_id': info.get('project_id', ''),
                'location': info.get('location', ''),
                'date': info.get('date', ''),
                'Mail Received - Date': info.get('summary_date', '')
            }
            basic_info_list.append(basic_info)
            
            # Add manufacturer stats with additional fields from basic info
            for stat in info.get('manufacturer_statistics', []):
                stat['file_name'] = os.path.basename(pdf_file)
                stat['company_name'] = info.get('company_name', '')
                stat['project_id'] = info.get('project_id', '')
                stat['location'] = info.get('location', '')
                stat['date'] = info.get('date', '')
                manufacturer_stats_list.append(stat)
                
        except Exception as e:
            print(f"Error processing {pdf_file}: {str(e)}")
    
    # Create DataFrames
    basic_df = pd.DataFrame(basic_info_list)
    manufacturer_df = pd.DataFrame(manufacturer_stats_list)
    
    # Check if we found any data before continuing
    if basic_df.empty and manufacturer_df.empty:
        print("No matching PDF files found. Make sure files contain 'Draft Report' or 'Draft Findings' in their names.")
        return None, None, None, None
    
    # Reorder columns in manufacturer_df to include all the new columns
    if not manufacturer_df.empty:
        column_order = [
            'file_name',
            'company_name',
            'project_id',
            'location',
            'date',
            'manufacturer_name',
            'stock_observation_percentage',
            'affected_loose_units',
            'affected_full_cases',
            'affected_loose_repeat_batch',
            'affected_cases_repeat_batch',
            'has_exponential'
        ]
        # Make sure all columns exist before reordering
        for col in column_order:
            if col not in manufacturer_df.columns:
                manufacturer_df[col] = ''
        manufacturer_df = manufacturer_df[column_order]
    
    # Reorder columns in basic_df to put Mail Received - Date after date
    if not basic_df.empty:
        column_order = ['file_name', 'company_name', 'project_id', 'location', 'date', 'Mail Received - Date']
        # Make sure all columns exist before reordering
        for col in column_order:
            if col not in basic_df.columns:
                if col == 'Mail Received - Date' and 'summary_date' in basic_df.columns:
                    # Rename summary_date to Mail Received - Date if it exists
                    basic_df = basic_df.rename(columns={'summary_date': 'Mail Received - Date'})
                else:
                    basic_df[col] = ''
        basic_df = basic_df[column_order]
    
    # Format the output filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    original_path = None
    report_path = None
    
    # 1. Save the original format Excel with Basic Info and Manufacturer Stats sheets
    if original_save_path:
        original_path = original_save_path
    else:
        original_filename = f'pdf_extraction_original_{timestamp}.xlsx'
        original_path = os.path.join(directory, original_filename)
    
    # Save the original format file
    with pd.ExcelWriter(original_path, engine='openpyxl') as writer:
        # Rename the summary_date column in basic_df to 'Mail Received - Date' before saving
        if 'summary_date' in basic_df.columns:
            basic_df = basic_df.rename(columns={'summary_date': 'Mail Received - Date'})
        
        basic_df.to_excel(writer, sheet_name='Basic Info', index=False)
        manufacturer_df.to_excel(writer, sheet_name='Manufacturer Stats', index=False)
    
    print(f"\nOriginal format saved to: {original_path}")
    print(f"  - 'Basic Info' sheet with {len(basic_df)} entries")
    print(f"  - 'Manufacturer Stats' sheet with {len(manufacturer_df)} entries")
    
    # 2. Now continue with the new format if report_save_path is provided
    if report_save_path:
        report_path = report_save_path
        
        # Prepare data for the Injured Tracker sheet
        injured_tracker_data = []
        
        # Define mapping between full names and short names for 'Injured Bottler'
        bottler_mapping = {
            'Moon Beverages Limited': 'Moon',
            'SLMG Beverages Private Limited': 'SLMG',
            'Enrich Agro Food Products Private Limited': 'Enrich',
            'Kandhari Beverages Limited': 'KBL',
            'Udaipur Beverages Limited': 'UBL',
            'Narmada Drinks Pvt Ltd': 'NDPL',
            'Ludhiana Beverages Private Limited': 'LBPL',
            'Kandhari Global Beverages Private Limited': 'KGB',
            'Superior Drinks Pvt. Ltd.': 'SDPL',
            'Hindustan Coca-Cola Beverages Pvt. Ltd.': 'HCCB',
            'Enrich Agro Food Products\nPrivate Limited': 'Enrich',
            'Enrich Agro Food Products Pvt. Ltd.': 'Enrich'
        }
        
        # Process basic_df to create injured_tracker_data
        if not basic_df.empty:
            for idx, row in basic_df.iterrows():
                # Determine the injured bottler from company_name
                injured_bottler = ''
                company_name = str(row['company_name']).strip() if not pd.isna(row['company_name']) else ''
                
                # Check if company_name matches any full name in the mapping
                for full_name, short_name in bottler_mapping.items():
                    if full_name.lower() in company_name.lower():
                        injured_bottler = short_name
                        break
                    
                    # Also check if the short name is in the project_id
                    if not injured_bottler and not pd.isna(row['project_id']):
                        project_id = str(row['project_id']).strip()
                        if short_name in project_id:
                            injured_bottler = short_name
                            break
                
                # Get location and survey no
                location = str(row['location']).strip() if not pd.isna(row['location']) else ''
                survey_no = str(row['project_id']).strip() if not pd.isna(row['project_id']) else ''
                
                # Create the injured mail subject by concatenating the three values with hyphens
                injured_mail_subject = ''
                if injured_bottler and location and survey_no:
                    injured_mail_subject = f"{injured_bottler}-{location}-{survey_no}"
                
                # Determine audit planned by
                audit_planned_by = 'BDO' if survey_no.startswith('SR') else 'EY'
                
                new_row = {
                    'Sr. No': idx + 1,
                    'Survey No': survey_no,
                    'Injured Bottler': injured_bottler,
                    'Location': location,
                    'Mail Received - Date': row['Mail Received - Date'],
                    'Report Received Date': row['date'],
                    'Injured Mail Subject': injured_mail_subject,
                    'Audit Planned by': audit_planned_by
                }
                injured_tracker_data.append(new_row)
        
        # Prepare data for the Source Report Summary sheet
        source_report_data = []
        project_count = {}
        
        # Process manufacturer_df to create source_report_data
        if not manufacturer_df.empty:
            for idx, row in manufacturer_df.iterrows():
                # Get project_id from the extracted data
                project_id = str(row['project_id']).strip() if not pd.isna(row['project_id']) else ''
                
                # Determine the sequential number for this project_id
                if project_id not in project_count:
                    project_count[project_id] = 1
                else:
                    project_count[project_id] += 1
                
                # Calculate TOTAL REPEATED
                loose_repeat = row['affected_loose_repeat_batch'] if not pd.isna(row['affected_loose_repeat_batch']) else 0
                cases_repeat = row['affected_cases_repeat_batch'] if not pd.isna(row['affected_cases_repeat_batch']) else 0
                total_repeated = loose_repeat + cases_repeat
                
                # If total_repeated is 0, use a hyphen
                total_repeated_display = '-' if total_repeated == 0 else total_repeated
                
                # Find the corresponding injured bottler from the basic info
                injured_bottler = ''
                for basic_row in basic_df.iterrows():
                    basic_project_id = str(basic_row[1]['project_id']).strip() if not pd.isna(basic_row[1]['project_id']) else ''
                    if basic_project_id == project_id:
                        company_name = str(basic_row[1]['company_name']).strip() if not pd.isna(basic_row[1]['company_name']) else ''
                        
                        # Try to match with bottler mapping for injured_bottler
                        for full_name, short_name in bottler_mapping.items():
                            if full_name.lower() in company_name.lower():
                                injured_bottler = short_name
                                break
                        
                        # If not found, check if any short name is in the project_id
                        if not injured_bottler:
                            for full_name, short_name in bottler_mapping.items():
                                if short_name in project_id:
                                    injured_bottler = short_name
                                    break
                        
                        break
                
                # Get manufacturer_name directly for Bottler Name column
                bottler_name = str(row['manufacturer_name']).strip() if not pd.isna(row['manufacturer_name']) else ''
                
                new_row = {
                    'No.': project_count[project_id],
                    'Survey No': project_id,
                    'Injured': injured_bottler,
                    'Location': row['location'],
                    'Bottler Name ': bottler_name,
                    'Report Rec. Date': row['date'],
                    '% of outlets at which stock was observed': row['stock_observation_percentage'],
                    'Loose no of units': row['affected_loose_units'],
                    'Full cases count': row['affected_full_cases'],
                    'Repeated LOOSE - Case': row['affected_loose_repeat_batch'],
                    'Repeated FULL - Case': row['affected_cases_repeat_batch'],
                    'TOTAL REPEATED': total_repeated_display,
                }
                source_report_data.append(new_row)
        
        # Convert to DataFrames
        injured_tracker_df = pd.DataFrame(injured_tracker_data)
        source_report_df = pd.DataFrame(source_report_data)
        
        # Save to Excel with the new sheet names
        try:
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                injured_tracker_df.to_excel(writer, sheet_name='Injured - Tracker 2022+2023', index=False)
                source_report_df.to_excel(writer, sheet_name='Source - Report Summary', index=False)
                
                # Format the worksheet
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Auto-adjust column width for all columns
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        
                        adjusted_width = (max_length + 2) * 1.2  # Add some padding
                        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 40)  # Cap width at 40
                    
                    # Center align Sr. No column
                    if sheet_name == 'Injured - Tracker 2022+2023':
                        for row in range(2, worksheet.max_row + 1):  # Start from row 2 (after header)
                            cell = worksheet['A' + str(row)]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                
            print(f"\nReport format saved to: {report_path}")
            print(f"  - 'Injured - Tracker 2022+2023' sheet with {len(injured_tracker_data)} entries")
            print(f"  - 'Source - Report Summary' sheet with {len(source_report_data)} entries")
            
        except Exception as e:
            print(f"Error saving to Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
    
    return basic_df, manufacturer_df, original_path, report_path


def update_final_tracker(extracted_file=None, final_tracker_file=None):
    """
    Update the final tracker Excel file with data from the extracted file.
    """
    # This function is obsolete and has been removed
    # The functionality has been replaced by creating a new Excel file
    # with the required sheet names directly in the process_all_pdfs_for_gui function
    pass


def process_all_pdfs(directory='.'):
    """Process all PDF files in the directory and its subdirectories and save results to Excel."""
    # Ask user to select a directory with PDFs
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    directory = filedialog.askdirectory(title="Select folder containing PDF files")
    if not directory:
        print("No directory selected. Operation cancelled.")
        return None, None
    
    # Use the new function to handle the PDF processing
    basic_df, manufacturer_df, _, _ = process_all_pdfs_for_gui(directory)
    return basic_df, manufacturer_df


def generate_report_from_data(basic_df, manufacturer_df, report_path):
    """
    Generate report format Excel from existing dataframes without reprocessing PDFs.
    
    Args:
        basic_df (DataFrame): Basic info dataframe with 'Basic Info' data
        manufacturer_df (DataFrame): Manufacturer stats dataframe with 'Manufacturer Stats' data
        report_path (str): Path where to save the report 
        
    Returns:
        tuple: (basic_df, manufacturer_df, None, report_path)
    """
    # Check if dataframes are empty
    if basic_df.empty and manufacturer_df.empty:
        print("No data found in the provided dataframes.")
        return None, None, None, None
    
    # Prepare data for the Injured Tracker sheet
    injured_tracker_data = []
    
    # Define mapping between full names and short names for 'Injured Bottler'
    bottler_mapping = {
        'Moon Beverages Limited': 'Moon',
        'SLMG Beverages Private Limited': 'SLMG',
        'Enrich Agro Food Products Private Limited': 'Enrich',
        'Kandhari Beverages Limited': 'KBL',
        'Udaipur Beverages Limited': 'UBL',
        'Narmada Drinks Pvt Ltd': 'NDPL',
        'Ludhiana Beverages Private Limited': 'LBPL',
        'Kandhari Global Beverages Private Limited': 'KGB',
        'Superior Drinks Pvt. Ltd.': 'SDPL',
        'Hindustan Coca-Cola Beverages Pvt. Ltd.': 'HCCB',
        'Enrich Agro Food Products\nPrivate Limited': 'Enrich',
        'Enrich Agro Food Products Pvt. Ltd.': 'Enrich'
    }
    
    # Process basic_df to create injured_tracker_data
    if not basic_df.empty:
        for idx, row in basic_df.iterrows():
            # Determine the injured bottler from company_name
            injured_bottler = ''
            company_name = str(row['company_name']).strip() if not pd.isna(row['company_name']) else ''
            
            # Check if company_name matches any full name in the mapping
            for full_name, short_name in bottler_mapping.items():
                if full_name.lower() in company_name.lower():
                    injured_bottler = short_name
                    break
                
                # Also check if the short name is in the project_id
                if not injured_bottler and not pd.isna(row['project_id']):
                    project_id = str(row['project_id']).strip()
                    if short_name in project_id:
                        injured_bottler = short_name
                        break
            
            # Get location and survey no
            location = str(row['location']).strip() if not pd.isna(row['location']) else ''
            survey_no = str(row['project_id']).strip() if not pd.isna(row['project_id']) else ''
            
            # Create the injured mail subject by concatenating the three values with hyphens
            injured_mail_subject = ''
            if injured_bottler and location and survey_no:
                injured_mail_subject = f"{injured_bottler}-{location}-{survey_no}"
            
            # Determine audit planned by
            audit_planned_by = 'BDO' if survey_no.startswith('SR') else 'EY'
            
            # Check if 'Mail Received - Date' or 'summary_date' column exists
            mail_date = None
            if 'Mail Received - Date' in row:
                mail_date = row['Mail Received - Date']
            elif 'summary_date' in row:
                mail_date = row['summary_date']
            
            new_row = {
                'Sr. No': idx + 1,
                'Survey No': survey_no,
                'Injured Bottler': injured_bottler,
                'Location': location,
                'Mail Received - Date': mail_date,
                'Report Received Date': row['date'],
                'Injured Mail Subject': injured_mail_subject,
                'Audit Planned by': audit_planned_by
            }
            injured_tracker_data.append(new_row)
    
    # Prepare data for the Source Report Summary sheet
    source_report_data = []
    project_count = {}
    
    # Process manufacturer_df to create source_report_data
    if not manufacturer_df.empty:
        for idx, row in manufacturer_df.iterrows():
            # Get project_id from the extracted data
            project_id = str(row['project_id']).strip() if not pd.isna(row['project_id']) else ''
            
            # Determine the sequential number for this project_id
            if project_id not in project_count:
                project_count[project_id] = 1
            else:
                project_count[project_id] += 1
            
            # Calculate TOTAL REPEATED
            loose_repeat = row['affected_loose_repeat_batch'] if not pd.isna(row['affected_loose_repeat_batch']) else 0
            cases_repeat = row['affected_cases_repeat_batch'] if not pd.isna(row['affected_cases_repeat_batch']) else 0
            total_repeated = loose_repeat + cases_repeat
            
            # If total_repeated is 0, use a hyphen
            total_repeated_display = '-' if total_repeated == 0 else total_repeated
            
            # Find the corresponding injured bottler from the basic info
            injured_bottler = ''
            for basic_row in basic_df.iterrows():
                basic_project_id = str(basic_row[1]['project_id']).strip() if not pd.isna(basic_row[1]['project_id']) else ''
                if basic_project_id == project_id:
                    company_name = str(basic_row[1]['company_name']).strip() if not pd.isna(basic_row[1]['company_name']) else ''
                    
                    # Try to match with bottler mapping for injured_bottler
                    for full_name, short_name in bottler_mapping.items():
                        if full_name.lower() in company_name.lower():
                            injured_bottler = short_name
                            break
                    
                    # If not found, check if any short name is in the project_id
                    if not injured_bottler:
                        for full_name, short_name in bottler_mapping.items():
                            if short_name in project_id:
                                injured_bottler = short_name
                                break
                    
                    break
            
            # Get manufacturer_name directly for Bottler Name column
            bottler_name = str(row['manufacturer_name']).strip() if not pd.isna(row['manufacturer_name']) else ''
            
            new_row = {
                'No.': project_count[project_id],
                'Survey No': project_id,
                'Injured': injured_bottler,
                'Location': row['location'],
                'Bottler Name ': bottler_name,
                'Report Rec. Date': row['date'],
                '% of outlets at which stock was observed': row['stock_observation_percentage'],
                'Loose no of units': row['affected_loose_units'],
                'Full cases count': row['affected_full_cases'],
                'Repeated LOOSE - Case': row['affected_loose_repeat_batch'],
                'Repeated FULL - Case': row['affected_cases_repeat_batch'],
                'TOTAL REPEATED': total_repeated_display,
            }
            source_report_data.append(new_row)
    
    # Convert to DataFrames
    injured_tracker_df = pd.DataFrame(injured_tracker_data)
    source_report_df = pd.DataFrame(source_report_data)
    
    # Save to Excel with the new sheet names
    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            injured_tracker_df.to_excel(writer, sheet_name='Injured - Tracker 2022+2023', index=False)
            source_report_df.to_excel(writer, sheet_name='Source - Report Summary', index=False)
            
            # Format the worksheet
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Auto-adjust column width for all columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    adjusted_width = (max_length + 2) * 1.2  # Add some padding
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 40)  # Cap width at 40
                
                # Center align Sr. No column
                if sheet_name == 'Injured - Tracker 2022+2023':
                    for row in range(2, worksheet.max_row + 1):  # Start from row 2 (after header)
                        cell = worksheet['A' + str(row)]
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
        print(f"\nReport format saved to: {report_path}")
        print(f"  - 'Injured - Tracker 2022+2023' sheet with {len(injured_tracker_data)} entries")
        print(f"  - 'Source - Report Summary' sheet with {len(source_report_data)} entries")
        
    except Exception as e:
        print(f"Error saving to Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return basic_df, manufacturer_df, None, None
    
    return basic_df, manufacturer_df, None, report_path


# Update the main function to reflect the changes
if __name__ == "__main__":
    print("PDF Extraction Tool")
    print("This tool extracts data from PDF files containing 'Draft Report' or 'Draft Findings' in their filenames")
    print("The extracted data will be saved to two Excel files:")
    print("  1. Original format with 'Basic Info' and 'Manufacturer Stats' sheets")
    print("  2. Report format with 'Injured - Tracker 2022+2023' and 'Source - Report Summary' sheets")
    print("\nAll dates will be standardized to DD-MM-YYYY format")
    print("\nSelect a folder containing PDF files to begin...")
    # Ask user to select a directory
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    directory = filedialog.askdirectory(title="Select folder containing PDF files")
    if directory:
        basic_df, manufacturer_df, original_path, _ = process_all_pdfs_for_gui(directory)
    else:
        print("No directory selected. Operation cancelled.")
